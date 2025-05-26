import pandas as pd

def filter_xls_rows(input_file, output_file, column_name, target_value):
    try:
        # Read the .xls file
        df = pd.read_excel(input_file)

        # Filter the rows where the column matches the target value
        filtered_df = df[df[column_name] == target_value]

        # Write the filtered rows to a new .xls file
        filtered_df.to_excel(output_file, index=False)

        print(f"Filtered data saved to '{output_file}' successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
input_xls = "Equity_Transaction_Summary_12-26may2025.xlsx"
output_xls = "filtered_data.xlsx"
column_to_check = "Stock Name"
value_to_match = "C D S L"

filter_xls_rows(input_xls, output_xls, column_to_check, value_to_match)


from openpyxl import load_workbook
from openpyxl.styles import PatternFill  

def highlight_rows(input_file, output_file, column_name, target_value):
    try:
        # Read the Excel file
        df = pd.read_excel(input_file)

        # Save a copy to output so we can style it
        df.to_excel(output_file, index=False)

        # Load the workbook and sheet with openpyxl
        wb = load_workbook(output_file)
        ws = wb.active

        # Find column index for the target column name
        header = [cell.value for cell in ws[1]]
        if column_name not in header:
            print(f"Column '{column_name}' not found.")
            return

        col_index = header.index(column_name) + 1  # 1-based indexing

        # Define red fill
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

        # Loop through rows and apply red fill if condition is met
        for row in ws.iter_rows(min_row=2):  # Skip header
            cell = row[col_index - 1]
            if cell.value == target_value:
                for c in row:
                    c.fill = red_fill

        # Save the modified file
        wb.save(output_file)
        print(f"Rows highlighted and saved to '{output_file}' successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
input_file = "filtered_data.xlsx"
output_file = "colored_filtered_data.xlsx"
column_name = "Buy/Sell"
target_value = "SELL"

highlight_rows(input_file, output_file, column_name, target_value)



#//NOT WORKING 
def highlight_conditional_buys(input_file, output_file):
    # Read data using pandas
    df = pd.read_excel(input_file)

    # Export to a temp output to work with openpyxl
    df.to_excel(output_file, index=False)

    # Load workbook for styling
    wb = load_workbook(output_file)
    ws = wb.active

    # Prepare yellow fill
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    # Identify column indexes (1-based for openpyxl)
    headers = [cell.value for cell in ws[1]]
    buy_sell_col = headers.index("Buy/Sell") + 1
    avg_cost_col = headers.index("Average Cost") + 1

    last_sell_avg_cost = None

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # skip header
        buy_sell = row[buy_sell_col - 1].value
        avg_cost = row[avg_cost_col - 1].value

        if buy_sell == "Sell" and avg_cost is not None:
            last_sell_avg_cost = avg_cost

        elif buy_sell == "Buy" and avg_cost is not None and last_sell_avg_cost is not None:
            if avg_cost > last_sell_avg_cost:
                # Highlight the entire row in yellow
                for cell in row:
                    cell.fill = yellow_fill

    wb.save(output_file)
    print(f"Highlighted rows saved in '{output_file}'.")

# Example usage
input_file = "colored_filtered_data.xlsx"
output_file = "colored_filtered_data_step2_highest_buy_highlighted.xlsx"

#highlight_conditional_buys(input_file, output_file) //NOT WORKING 
